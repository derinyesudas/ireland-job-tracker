"""
Build data/companies.json - the list of companies we watch, and where to watch them.

Two sources are merged:

  1. THE OFFICIAL SPONSOR REGISTER
     The Department of Enterprise, Tourism and Employment publishes a
     spreadsheet of every company issued an employment permit this year:
       https://enterprise.gov.ie/en/publications/publication-files/
         employment-permits-issued-to-companies-2026.xlsx
     This is the single most reliable answer to "who actually sponsors?" -
     it is not a claim on a job board, it is a record of permits granted.
     The permit count also tells us HOW active a sponsor they are.

  2. A CURATED LIST of large Irish graduate employers (data/seed_companies.json)
     Some big employers hire graduates heavily without appearing high on the
     permit register, so they are added by hand.

For each company we then work out HOW they hire, by probing the public job
feeds of the major applicant-tracking systems with slug variations of the
company name. A company is only kept if a probe actually returns a live board -
nothing is guessed or invented.

Run weekly (discovery is slow); the 15-minute scraper just reads the output.

    python scripts/build_companies.py --limit 400
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from scraper.ats_clients import FETCHERS, FetchError, USER_AGENT  # noqa: E402

DATA = ROOT / "data"
PERMITS_URL = (
    "https://enterprise.gov.ie/en/publications/publication-files/"
    "employment-permits-issued-to-companies-2026.xlsx"
)
PERMITS_CACHE = DATA / "permits_raw.json"
SEED_FILE = DATA / "seed_companies.json"
OUT_FILE = DATA / "companies.json"
DISCOVERY_CACHE = DATA / "ats_discovery_cache.json"

# Probed in this order - cheapest and most common first. These are the
# platforms whose board token can be guessed from a company name; Workday and
# Oracle need a tenant and site id that only the careers page reveals, so they
# come in through scripts/resolve_careers.py instead.
PROBE_ORDER = [
    "greenhouse", "lever", "workable", "smartrecruiters",
    "ashby", "recruitee", "pinpoint", "personio",
]

LEGAL_SUFFIXES = re.compile(
    r"\b(limited|ltd\.?|plc|dac|ulc|teoranta|teo\.?|unlimited company|"
    r"holdings?|group|international|ireland|\(ireland\)|eire|company|"
    r"services|solutions|technologies|technology|global|europe|emea|"
    r"operations|management|consulting|partners|associates|inc\.?|corp\.?|"
    r"corporation|llc|llp|gmbh|b\.?v\.?|s\.?a\.?)\b",
    re.I,
)
NON_ALNUM = re.compile(r"[^a-z0-9]+")


def slug_variants(name: str) -> list[str]:
    """Generate plausible board tokens for a company name, best guess first."""
    base = name.lower().strip()
    base = re.sub(r"[\"'’]", "", base)

    stripped = LEGAL_SUFFIXES.sub(" ", base)
    stripped = re.sub(r"\s+", " ", stripped).strip()

    candidates: list[str] = []

    def push(value: str) -> None:
        value = NON_ALNUM.sub("", value)
        if 2 < len(value) <= 40 and value not in candidates:
            candidates.append(value)

    def push_hyphenated(value: str) -> None:
        # Same sanitising as push(), but keeping the hyphens - board tokens
        # come in both shapes. Without stripping punctuation first this
        # produced tokens like "citco-fund-(-)".
        value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
        value = re.sub(r"-{2,}", "-", value)
        if 2 < len(value) <= 40 and value not in candidates:
            candidates.append(value)

    push(stripped)                      # "accenture"
    push(base)                          # "accentureireland"
    if " " in stripped:
        push(stripped.split(" ")[0])    # first word only
    push_hyphenated(stripped)           # "citco-fund"

    return candidates[:4]


# --------------------------------------------------------------- permit sheet

def download_permit_register() -> list[dict]:
    """Download and parse the government permit spreadsheet."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed - run: pip install openpyxl")
        return []

    print(f"downloading {PERMITS_URL}")
    req = urllib.request.Request(PERMITS_URL, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            blob = resp.read()
    except (urllib.error.URLError, OSError) as exc:
        print(f"  could not download permit register: {exc}")
        if PERMITS_CACHE.exists():
            print("  using cached copy")
            return json.loads(PERMITS_CACHE.read_text(encoding="utf-8"))
        return []

    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet in wb.worksheets:
        header_row = None
        name_col = count_col = None

        for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True)):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            for c_idx, cell in enumerate(cells):
                if any(k in cell for k in ("employer", "company", "organisation", "name")):
                    name_col = c_idx
                    header_row = r_idx + 1
                if any(k in cell for k in ("issued", "total", "number", "count", "permits")):
                    count_col = c_idx
            if name_col is not None:
                break

        if name_col is None:
            continue

        for row in sheet.iter_rows(min_row=(header_row or 1) + 1, values_only=True):
            if not row or name_col >= len(row):
                continue
            raw_name = row[name_col]
            if not raw_name or not str(raw_name).strip():
                continue
            name = str(raw_name).strip()
            if name.lower() in ("total", "grand total", "employer", "company name"):
                continue
            permits = 1
            if count_col is not None and count_col < len(row):
                try:
                    permits = int(float(row[count_col] or 1))
                except (TypeError, ValueError):
                    permits = 1
            rows.append({"name": name, "permits": permits})

    # Merge duplicate company rows
    merged: dict[str, int] = {}
    for row in rows:
        key = row["name"]
        merged[key] = merged.get(key, 0) + row["permits"]

    out = [{"name": k, "permits": v} for k, v in merged.items()]
    out.sort(key=lambda r: -r["permits"])
    print(f"  parsed {len(out)} sponsoring companies from the register")

    DATA.mkdir(exist_ok=True)
    PERMITS_CACHE.write_text(json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")
    return out


def sponsor_tier(permits: int) -> int:
    if permits >= 25:
        return 3
    if permits >= 5:
        return 2
    if permits >= 1:
        return 1
    return 0


# ------------------------------------------------------------- ats discovery

# Discovery only needs to know whether a board EXISTS, not what is on it.
# The full readers fetch each job's description one page at a time, which is
# right for scraping and ruinous for probing tens of thousands of candidate
# slugs - so existence is checked with one cheap request instead.
QUICK_PROBE = {
    "greenhouse": "https://boards-api.greenhouse.io/v1/boards/{t}/jobs",
    "lever": "https://api.lever.co/v0/postings/{t}?mode=json&limit=1",
    "ashby": "https://api.ashbyhq.com/posting-api/job-board/{t}",
    "workable": "https://apply.workable.com/api/v1/widget/accounts/{t}",
    "smartrecruiters": "https://api.smartrecruiters.com/v1/companies/{t}/postings?limit=1",
    "recruitee": "https://{t}.recruitee.com/api/offers/",
    "pinpoint": "https://{t}.pinpointhq.com/postings.json",
}


def _count_in(payload) -> int:
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict):
        for key in ("jobs", "content", "offers", "data", "postings", "results"):
            value = payload.get(key)
            if isinstance(value, list):
                return len(value)
        if payload.get("totalFound"):
            try:
                return int(payload["totalFound"])
            except (TypeError, ValueError):
                pass
    return 0


def probe(ats: str, token: str) -> int:
    """Return a job count if this board exists and has postings, else -1."""
    from scraper.ats_clients import _get

    template = QUICK_PROBE.get(ats)
    if template:
        try:
            count = _count_in(_get(template.format(t=token)))
        except (FetchError, Exception):  # noqa: BLE001
            return -1
        return count if count > 0 else -1

    fetcher = FETCHERS.get(ats)
    if not fetcher:
        return -1
    try:
        jobs = fetcher(token)
    except (FetchError, Exception):  # noqa: BLE001
        return -1
    return len(jobs) if jobs else -1


def discover(company: dict, cache: dict) -> dict | None:
    """Find which ATS a company uses. Returns an enriched record or None."""
    name = company["name"]
    if name in cache:
        cached = cache[name]
        if cached is None:
            return None
        return {**company, **cached}

    for token in slug_variants(name):
        for ats in PROBE_ORDER:
            count = probe(ats, token)
            if count > 0:
                found = {"ats": ats, "token": token, "jobs_at_discovery": count}
                cache[name] = found
                print(f"  FOUND  {name:<45} -> {ats}/{token} ({count} jobs)")
                return {**company, **found}
            time.sleep(0.05)

    cache[name] = None
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=400,
                    help="how many top sponsors to probe (they are slow)")
    ap.add_argument("--min-permits", type=int, default=2,
                    help="ignore companies with fewer permits than this")
    ap.add_argument("--skip-download", action="store_true")
    ap.add_argument("--cap", type=int, default=290,
                    help="never grow the tracker beyond this many companies")
    args = ap.parse_args()

    DATA.mkdir(exist_ok=True)

    # 1. sponsors from the government register
    if args.skip_download and PERMITS_CACHE.exists():
        sponsors = json.loads(PERMITS_CACHE.read_text(encoding="utf-8"))
    else:
        sponsors = download_permit_register()

    sponsors = [s for s in sponsors if s["permits"] >= args.min_permits]
    sponsors = sponsors[: args.limit]

    # 2. curated graduate employers
    seed = []
    if SEED_FILE.exists():
        seed = json.loads(SEED_FILE.read_text(encoding="utf-8"))
    print(f"{len(sponsors)} sponsors to probe + {len(seed)} curated employers")

    permits_by_name = {s["name"].lower(): s["permits"] for s in sponsors}

    cache = {}
    if DISCOVERY_CACHE.exists():
        cache = json.loads(DISCOVERY_CACHE.read_text(encoding="utf-8"))

    results: list[dict] = []

    # Curated entries already carry a known ats + token, so just verify them.
    for entry in seed:
        count = probe(entry["ats"], entry["token"])
        if count > 0:
            entry = dict(entry)
            entry["jobs_at_discovery"] = count
            entry["priority"] = True
            key = entry["name"].lower()
            entry["permits"] = permits_by_name.get(key, entry.get("permits", 0))
            entry["sponsor_tier"] = sponsor_tier(entry["permits"])
            results.append(entry)
            print(f"  OK     {entry['name']:<45} -> {entry['ats']}/{entry['token']} ({count} jobs)")
        else:
            print(f"  dead   {entry['name']:<45} -> {entry['ats']}/{entry['token']}")

    known_names = {r["name"].lower() for r in results}

    # Probe the sponsor register
    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {
            pool.submit(discover, s, cache): s
            for s in sponsors
            if s["name"].lower() not in known_names
        }
        for fut in as_completed(futures):
            found = fut.result()
            if found:
                found["sponsor_tier"] = sponsor_tier(found.get("permits", 0))
                found["priority"] = found.get("permits", 0) >= 25
                results.append(found)

    DISCOVERY_CACHE.write_text(json.dumps(cache, indent=1, ensure_ascii=False), encoding="utf-8")

    # MERGE with whatever is already tracked, never replace it.
    #
    # This script used to overwrite companies.json outright. Running it would
    # have silently destroyed every company the careers-site resolver had
    # worked out - which is most of them, and the ones with the best jobs.
    # Anything already in the file stays; the permit register only adds.
    by_key: dict[tuple, dict] = {}
    if OUT_FILE.exists():
        try:
            for c in json.loads(OUT_FILE.read_text(encoding="utf-8")):
                by_key[(c.get("ats"), c.get("token"))] = c
        except json.JSONDecodeError:
            pass
    kept_existing = len(by_key)

    # The tracker is deliberately capped. Watching every sponsor in the country
    # would be slower, noisier and no more useful: a company earns a place by
    # being a heavy permit sponsor, not by existing. Companies already tracked
    # are never displaced - the cap only limits what gets added on top.
    added = 0
    for r in sorted(results, key=lambda x: -x.get("permits", 0)):
        key = (r.get("ats"), r.get("token"))
        if key not in by_key and args.cap and len(by_key) >= args.cap:
            continue
        if key in by_key:
            # Same feed found again - only fill in the permit evidence, and
            # never downgrade a sponsor tier the register already established.
            existing = by_key[key]
            if r.get("permits", 0) > existing.get("permits", 0):
                existing["permits"] = r["permits"]
                existing["sponsor_tier"] = max(
                    existing.get("sponsor_tier", 0), r.get("sponsor_tier", 0)
                )
        else:
            by_key[key] = r
            added += 1

    final = sorted(by_key.values(),
                   key=lambda c: (-c.get("fit_rank", 0), -c.get("permits", 0)))
    print(f"\nkept {kept_existing} already-tracked companies, added {added} new ones"
          + (f" (cap {args.cap})" if args.cap else ""))

    OUT_FILE.write_text(json.dumps(final, indent=1, ensure_ascii=False), encoding="utf-8")

    priority = sum(1 for r in final if r.get("priority"))
    print(f"\nwrote {len(final)} companies to {OUT_FILE} ({priority} priority)")
    by_ats: dict[str, int] = {}
    for r in final:
        by_ats[r["ats"]] = by_ats.get(r["ats"], 0) + 1
    print(f"by platform: {by_ats}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
